import logging
from io import BytesIO
from datetime import datetime
from typing import Optional
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from app.services.inventory_service import inventory_service
from app.services.alert_service import alert_service

logger = logging.getLogger(__name__)


class ExportService:
    """Service for exporting data to Excel"""
    
    async def export_current_inventory(self) -> bytes:
        """Export current inventory to Excel"""
        inventory = await inventory_service.get_current_inventory()
        
        data = []
        for bin_data in inventory:
            data.append({
                "Bin ID": bin_data.bin_id,
                "Location": f"Row {bin_data.row}, Position {bin_data.position}",
                "Article Type": bin_data.article_type,
                "Article Name": bin_data.article_name,
                "Current Quantity": bin_data.current_quantity,
                "Max Capacity": bin_data.max_capacity,
                "Fill %": bin_data.fill_percentage,
                "Status": bin_data.status.value.upper(),
                "Last Updated": bin_data.last_updated
            })
        
        return self._create_excel(data, "Current Inventory")
    
    async def export_historical_data(
        self,
        start_date: str,
        end_date: str,
        bin_ids: Optional[list[str]] = None
    ) -> bytes:
        """Export historical data to Excel"""
        all_data = await inventory_service.get_all_historical_data(start_date, end_date)
        
        # Filter by bin IDs if provided
        if bin_ids:
            all_data = [d for d in all_data if d["bin_id"] in bin_ids]
        
        # Flatten data for export
        data = []
        for bin_data in all_data:
            for point in bin_data["data"]:
                data.append({
                    "Bin ID": bin_data["bin_id"],
                    "Timestamp": point["timestamp"],
                    "Quantity": point["quantity"],
                    "Weight (g)": point["weight_grams"]
                })
        
        return self._create_excel(data, "Historical Data", {
            "Date Range": f"{start_date} to {end_date}",
            "Bins": ", ".join(bin_ids) if bin_ids else "All bins"
        })
    
    async def export_alerts(
        self,
        start_date: Optional[str] = None,
        end_date: Optional[str] = None,
        include_acknowledged: bool = True
    ) -> bytes:
        """Export alerts to Excel"""
        alerts, _ = await alert_service.get_alert_history(1, 10000)
        
        # Filter by date range
        if start_date:
            alerts = [a for a in alerts if a.created_at >= start_date]
        if end_date:
            alerts = [a for a in alerts if a.created_at <= end_date]
        if not include_acknowledged:
            alerts = [a for a in alerts if not a.is_acknowledged]
        
        data = []
        for alert in alerts:
            data.append({
                "Alert ID": alert.id,
                "Bin ID": alert.bin_id,
                "Alert Type": alert.alert_type,
                "Message": alert.message,
                "Quantity at Alert": alert.quantity_at_alert,
                "Threshold": alert.threshold_value,
                "Acknowledged": "Yes" if alert.is_acknowledged else "No",
                "Acknowledged At": alert.acknowledged_at or "",
                "Acknowledged By": alert.acknowledged_by or "",
                "Created At": alert.created_at
            })
        
        return self._create_excel(data, "Alerts")
    
    async def export_summary_report(self) -> bytes:
        """Export comprehensive summary report"""
        inventory = await inventory_service.get_current_inventory()
        summary = await inventory_service.get_inventory_summary()
        active_alerts = await alert_service.get_active_alerts()
        
        wb = Workbook()
        
        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        summary_data = [
            ("Metric", "Value"),
            ("Total Bins", summary.total_bins),
            ("Normal Stock", summary.normal_count),
            ("Low Stock", summary.low_count),
            ("Critical Stock", summary.critical_count),
            ("Empty Bins", summary.empty_count),
            ("Total Items", summary.total_items),
            ("Active Alerts", summary.alerts_active),
            ("Report Generated", datetime.now().isoformat())
        ]
        
        for row_idx, (key, value) in enumerate(summary_data, 1):
            ws_summary.cell(row=row_idx, column=1, value=key)
            ws_summary.cell(row=row_idx, column=2, value=value)
        
        # Inventory sheet
        ws_inventory = wb.create_sheet("Inventory")
        
        headers = ["Bin ID", "Location", "Article", "Type", "Quantity", 
                   "Max Capacity", "Fill %", "Status", "Last Updated"]
        for col_idx, header in enumerate(headers, 1):
            ws_inventory.cell(row=1, column=col_idx, value=header)
        
        for row_idx, bin_data in enumerate(inventory, 2):
            ws_inventory.cell(row=row_idx, column=1, value=bin_data.bin_id)
            ws_inventory.cell(row=row_idx, column=2, value=f"Row {bin_data.row}, Position {bin_data.position}")
            ws_inventory.cell(row=row_idx, column=3, value=bin_data.article_name)
            ws_inventory.cell(row=row_idx, column=4, value=bin_data.article_type)
            ws_inventory.cell(row=row_idx, column=5, value=bin_data.current_quantity)
            ws_inventory.cell(row=row_idx, column=6, value=bin_data.max_capacity)
            ws_inventory.cell(row=row_idx, column=7, value=bin_data.fill_percentage)
            ws_inventory.cell(row=row_idx, column=8, value=bin_data.status.value.upper())
            ws_inventory.cell(row=row_idx, column=9, value=bin_data.last_updated)
        
        # Active Alerts sheet
        if active_alerts:
            ws_alerts = wb.create_sheet("Active Alerts")
            
            alert_headers = ["Bin ID", "Type", "Message", "Quantity", "Threshold", "Created"]
            for col_idx, header in enumerate(alert_headers, 1):
                ws_alerts.cell(row=1, column=col_idx, value=header)
            
            for row_idx, alert in enumerate(active_alerts, 2):
                ws_alerts.cell(row=row_idx, column=1, value=alert.bin_id)
                ws_alerts.cell(row=row_idx, column=2, value=alert.alert_type)
                ws_alerts.cell(row=row_idx, column=3, value=alert.message)
                ws_alerts.cell(row=row_idx, column=4, value=alert.quantity_at_alert)
                ws_alerts.cell(row=row_idx, column=5, value=alert.threshold_value)
                ws_alerts.cell(row=row_idx, column=6, value=alert.created_at)
        
        # Auto-size columns for all sheets
        for ws in wb.worksheets:
            for column_cells in ws.columns:
                max_length = 0
                column = column_cells[0].column_letter
                for cell in column_cells:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                ws.column_dimensions[column].width = min(max_length + 2, 50)
        
        # Save to buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        logger.info("Summary report generated")
        return buffer.getvalue()
    
    def _create_excel(
        self,
        data: list[dict],
        sheet_name: str,
        metadata: Optional[dict] = None
    ) -> bytes:
        """Create Excel file from data"""
        wb = Workbook()
        
        # Add metadata sheet if provided
        if metadata:
            ws_meta = wb.active
            ws_meta.title = "Export Info"
            
            row_idx = 1
            for key, value in metadata.items():
                ws_meta.cell(row=row_idx, column=1, value=key)
                ws_meta.cell(row=row_idx, column=2, value=value)
                row_idx += 1
            
            ws_meta.cell(row=row_idx, column=1, value="Generated At")
            ws_meta.cell(row=row_idx, column=2, value=datetime.now().isoformat())
            
            ws_data = wb.create_sheet(sheet_name)
        else:
            ws_data = wb.active
            ws_data.title = sheet_name
        
        # Write data
        if data:
            headers = list(data[0].keys())
            for col_idx, header in enumerate(headers, 1):
                ws_data.cell(row=1, column=col_idx, value=header)
            
            for row_idx, row_data in enumerate(data, 2):
                for col_idx, header in enumerate(headers, 1):
                    ws_data.cell(row=row_idx, column=col_idx, value=row_data.get(header))
            
            # Auto-size columns
            for col_idx, header in enumerate(headers, 1):
                max_length = len(header)
                for row_data in data:
                    cell_value = str(row_data.get(header, ""))
                    if len(cell_value) > max_length:
                        max_length = len(cell_value)
                ws_data.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 50)
        
        # Save to buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        logger.info(f"Excel export generated: {sheet_name} with {len(data)} rows")
        return buffer.getvalue()


# Singleton instance
export_service = ExportService()
